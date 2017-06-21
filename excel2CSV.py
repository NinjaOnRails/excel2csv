#! python3

import openpyxl, os, csv

for excelFile in os.listdir('.'):
    # Skip non-xlsx files, load the workbook object.
    if not excelFile.endswith('.xlsx'):
        continue
    wb = openpyxl.load_workbook(excelFile)
    for sheetName in wb.get_sheet_names():
        # Loop through every sheet in the workbook.
        sheet = wb.get_sheet_by_name(sheetName)

        # Create the CSV filename from the Excel filename and sheet title.
        CSVname = excelFile[:-5] + sheetName +'.csv'
        # Create the csv.writer object for this CSV file.
        csvFileObj = open(CSVname, 'w', newline='')
        csvWriter = csv.writer(csvFileObj)

        # Loop through every row in the sheet.
        for rowNum in range(1, sheet.max_row + 1):
            rowData = []    # append each cell to this list
            # Loop through each cell in the row.
            for colNum in range(1, sheet.max_column + 1):
                # Append each cell's data to rowData.
                rowData.append(sheet.cell(row=rowNum, column=colNum).value)

            # Write the rowData list to the CSV file.
            csvWriter.writerow(rowData)
        csvFileObj.close()
